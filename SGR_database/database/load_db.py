import re

import pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from openpyxl import load_workbook


class IndustryToDB(object):
    def __init__(self, excel_file):
        self.excel_file = excel_file
        (
            self.industry_group_1,
            self.industry_group_2,
            self.industry,
        ) = self.get_industries()

    def get_industries(self):
        wb = load_workbook(self.excel_file)
        try:
            ws = wb["Bản đồ ngành"]
        except IndexError:
            raise IndexError("There is no such sheet")
        else:
            industry_group_1 = set()
            industry_group_2 = set()
            industry = dict()
            # to skip the heading
            g = ws.iter_rows(values_only=True)
            next(g)

            for couple in g:
                industry_2, industry_1 = couple
                industry_group_1.add(industry_1)
                industry_group_2.add(industry_2)
                industry[industry_2] = industry_1
            return industry_group_1, industry_group_2, industry

    def industry_level1_to_db(self, IndustryLevel1):
        """
        input: IndustryLevel1 model
        output: create new object to IndustryLevel1 model
        """

        industries = list(self.industry_group_1)
        for industry in industries:
            IndustryLevel1.objects.create(industry_level_1=industry)

    def industry_to_db(self, IndustryLevel1, Industry):
        """
        input: IndustryLevel1 model, Industry model
        output: create new object to Industry model
        """

        for key, value in self.industry.items():
            level_1 = IndustryLevel1.objects.get(industry_level_1=value)
            Industry.objects.create(industry_level_1=level_1, industry_level_2=key)


class CompanyToDB(object):
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.stock_exchange_dict = self.get_companies()

    def get_companies(self):
        """
        output: return dictionary {'EXCHANGE':{'IndustryLevel2':[company name, company code]}}
        """

        wb = load_workbook(filename=self.excel_file)
        ws = wb["Danh sách tất cả công ty"]
        stock_exchange_dict = dict()

        # to skip the heading
        g = ws.iter_rows(values_only=True)
        next(g)
        # -------------------#

        for row in g:
            try:
                (
                    _,
                    code_vs,
                    company_name,
                    industry_level_1,
                    industry_level_2,
                    exchange,
                    issue_amount,
                ) = row
            except ValueError:
                print(row)
                raise ValueError("this is the bad one")
            if exchange not in stock_exchange_dict:
                stock_exchange_dict[exchange] = {}
            if industry_level_2 not in stock_exchange_dict[exchange]:
                stock_exchange_dict[exchange][industry_level_2] = []
            stock_exchange_dict[exchange][industry_level_2].append(
                (
                    company_name,
                    code_vs,
                )
            )
        return stock_exchange_dict

    def company_to_db(self, Company, StockExchange, Industry):
        """
        input: Company model, StockExchange model, Industry model
        output: create object for Company model
        """

        for stock_exchange, value_1 in self.stock_exchange_dict.items():
            exchange = StockExchange.objects.get(stock_exchange=stock_exchange)
            for industry_level_2, value_2 in value_1.items():
                try:
                    industry = Industry.objects.get(industry_level_2=industry_level_2)
                    for value_3 in value_2:
                        name, code_vs = value_3
                        Company.objects.create(
                            code_vs=code_vs,
                            name=name,
                            industry=industry,
                            exchange=exchange,
                        )
                except ObjectDoesNotExist as e:
                    print(industry_level_2, value_2)
                    raise e


class MatchCode(object):
    def __init__(self, vs_file, fiin_file):
        self.vs_file = vs_file
        self.fiin_file = fiin_file
        self.fiin = self.get_fiin_df()
        self.vs = self.get_vs_df()
        self.fiin_dict = self.get_fiin_dict()
        self.vs_dict = self.get_vs_dict()
        self.fiin_diff, self.vs_diff, self.shared = self.get_diff()
        self.matches = self.match_effort()
        self.ticket_dict = self.prepare_ticket_dict()

    def get_fiin_df(self):
        fiin = pd.read_excel(
            self.fiin_file, header=8, index_col=0, sheet_name="ExportData"
        )
        try:
            assert fiin.columns[0] == "Ticker"
            assert type(fiin.index[-15]) == float
        except AssertionError as e:
            raise e(
                "The data structure is not as expected, please check the difference"
            )
        else:
            fiin.drop(fiin.tail(15).index, inplace=True)
            return fiin

    def get_vs_df(self):
        vs = pd.read_excel(
            self.vs_file, header=0, index_col=0, sheet_name="Danh sách tất cả công ty"
        )
        try:
            assert vs.columns[0] == "Mã"
        except AssertionError as e:
            raise e('Expect "Mã" as the first column header')
        else:
            return vs

    def get_fiin_dict(self):
        """
        output: dictionary {'Ticker': 'OrganShortName'}
        """

        df = self.fiin.copy()
        df.set_index("Ticker", inplace=True)
        return df.to_dict()["OrganShortName"]

    def get_vs_dict(self):
        """
        output: dictionary {'Mã': 'Tên công ty'}
        """

        df = self.vs.copy()
        df.set_index("Mã", inplace=True)
        return df.to_dict()["Tên công ty"]

    def get_diff(self):
        """
        output:
            fiin_diff: list of (company code, company name) for companies that are only in fiin dataset
            vs_diff: list of (company code, company name) for companies that are only in vs dataset
            shared: list of (company code, company name) for companies that are in both dataset
        """

        fiin_tickers = set(self.fiin_dict.keys())
        vs_tickers = set(self.vs_dict.keys())

        fiin_diff = [
            (ticker, self.fiin_dict[ticker]) for ticker in (fiin_tickers - vs_tickers)
        ]
        vs_diff = [
            (ticker, self.vs_dict[ticker]) for ticker in (vs_tickers - fiin_tickers)
        ]
        shared = [
            (ticker, self.fiin_dict[ticker]) for ticker in (vs_tickers & fiin_tickers)
        ]

        return (fiin_diff, vs_diff, shared)

    def match_effort(self):
        """
        output: dictionary of match fiin company name to vs company name
        """

        # Try to match the different companies from fiin to vs dataset
        matches = []
        for fiin_com in self.fiin_diff:
            temp = []
            for vs_com in self.vs_diff:
                q = re.compile(r".*{}.*".format(fiin_com[1]), re.IGNORECASE)
                m = q.match(vs_com[1])
                if m:
                    temp.append((fiin_com, vs_com))

            if len(temp) == 1:  # there is only 1 match
                matches.append((fiin_com, vs_com))
            elif len(temp) > 1:  # there are more than 1 match
                print(temp)
        return matches

    def prepare_ticket_dict(self):
        """
        return the final dictionary: {vs_code: fiin_code}
        """
        ticket_dict = dict()
        for company in self.shared:
            ticket_dict[company[0]] = company[0]

        for match in self.matches:
            ticket_dict[match[1][0]] = match[0][0]

        return ticket_dict

    def add_fiin_code(self, Company):
        """
        input: Company model
        output: add fiin_code to company objects in Company model
        """

        for code_vs in self.ticket_dict:
            company = Company.objects.get(code_vs=code_vs)
            company.code_fiin = self.ticket_dict[code_vs]
            company.save()


class ReportType(object):
    """
    create statement objects for Statement model
    """

    REPORTS = [
        "Cân đối kế toán",
        "Kết quả Kinh doanh",
        "Lưu chuyển tiền tệ - Gián tiếp",
        "Lưu chuyển tiền tệ - Trực tiếp",
        "Thuyết minh",
        "Chỉ số tài chính",
        "Kết quả kiểm toán",
    ]

    def to_report_db(self, Statement, statement_category=None):
        """
        input:
            Statement model,
            statement_category: category in Statement.Categories.choices
                                default == None
        output:
            add new statement in to Statement model
        """

        for report in self.REPORTS:
            # check IntegrityError for duplicate report in the same category
            try:
                if statement_category:
                    # check the statement_category input is in the list of allowed categories
                    if statement_category in Statement.Categories:
                        Statement.objects.create(
                            statement_name=report, statement_category=statement_category
                        )
                    else:
                        raise Exception(
                            f"statement_category {statement_category} is not in categories list"
                        )
                else:
                    Statement.objects.create(statement_name=report)

            except IntegrityError:
                print(f"statement '{report}' already existed")
                pass


class ReportRow(object):
    def __init__(self, file, report_category="DN"):
        self.file = file
        self.report_category = report_category
        self.reports = self.get_reports()

    def get_reports(self):
        """
        output: reports dictionary contain: {report name: [report rows]}
        """

        wb = load_workbook(filename=self.file)
        ws = wb[self.report_category]

        reports = dict()
        for col in ws.iter_cols(values_only=True):
            # ignore value that is none, which is not string type, accept everything that is string
            reports[col[0]] = [value for value in col[1:] if type(value) == str]
        return reports

    def report_to_db(self, Statement, StatementRow):
        """
        input: Statement model, StatementRow model
        output: create object for StatementRow model
        Problem: Have not catch Integrity Error
        """

        q = re.compile(r"^(\s*)([\w(].*$)")

        for key, value in self.reports.items():
            try:
                statement = Statement.objects.get(
                    statement_name=key, statement_category=self.report_category
                )
                row_order = 1
            except ObjectDoesNotExist as e:
                print(
                    f"cannot retrieve object that has: {key} and {self.report_category}"
                )
                raise e

            for row in value:
                m = q.match(row)
                row_properties = ""
                try:
                    # try to add properties to each row
                    # There are two levels of indent is 5 or 10 in the dataset
                    if len(m.group(1)) == 5:
                        row_properties += "indented_1,"
                    if len(m.group(1)) == 10:
                        row_properties += "indented_2,"
                    # capture CAPITALIEZD row
                    if m.group(2).isupper():
                        row_properties += "bold,"

                    StatementRow.objects.create(
                        statement=statement,
                        title=row.strip(),
                        row_properties=row_properties,
                        row_order=row_order,
                    )
                    row_order += 1
                except AttributeError as e:
                    # catch AttributeError when m is not match
                    print(f"this row: {value} in statement {key} is invalid")
                    raise e
                except IntegrityError as e:
                    print(f"{row_order}: {statement}: {row}  has already exist")
                    raise e
