from openpyxl import load_workbook
import os
import re

wb = load_workbook(filename="database/excel/statement_rows.xlsx")
ws = wb["DN"]

q = re.compile(r"^(\s*)([\w(].*$)")
reports = {}
for col in ws.iter_cols(values_only=True):
    reports[col[0]] = [value for value in col[1:] if type(value) == str]

for row in reports["Kết quả kiểm toán"]:
    m = q.match(row)
    try:  # catch AttributeError when m is not match
        if m.group(1):
            print(len(m.group(1)))
            print(m.group(0))
    except AttributeError as e:
        print(f"this row: {row} in statement is invalid")
        raise e
print("done")
