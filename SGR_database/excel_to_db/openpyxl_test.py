import re

import pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import load_workbook


class IndustryToDB(object):
    def __init__(self, excel_file):
        self.excel_file = excel_file
        (
            self.industry_group_1,
            self.industry_group_2,
            self.industry,
        ) = self.get_industries()

    def get_industries(self):
        wb = load_workbook(self.excel_file)
        try:
            ws = wb["Bản đồ ngành"]
        except IndexError:
            raise IndexError("There is no such sheet")
        else:
            industry_group_1 = set()
            industry_group_2 = set()
            industry = dict()
            # to skip the heading
            g = ws.iter_rows(values_only=True)
            next(g)
            # -------------------#
            for couple in g:
                industry_2, industry_1 = couple
                industry_group_1.add(industry_1)
                industry_group_2.add(industry_2)
                industry[industry_2] = industry_1
            return industry_group_1, industry_group_2, industry

    def industry_level1_to_db(self, IndustryLevel1):
        """
        input: IndustryLevel1 model
        output: create new object to IndustryLevel1 model
        """

        industries = list(self.industry_group_1)
        for industry in industries:
            IndustryLevel1.objects.create(industry_level_1=industry)

    def industry_to_db(self, IndustryLevel1, Industry):
        """
        input: IndustryLevel1 model, Industry model
        output: create new object to Industry model
        """

        for key, value in self.industry.items():
            level_1 = IndustryLevel1.objects.get(industry_level_1=value)
            Industry.objects.create(industry_level_1=level_1, industry_level_2=key)


class CompanyToDB(object):
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.stock_exchange_dict = self.get_companies()

    def get_companies(self):
        """
        output: return dictionary {'EXCHANGE':{'IndustryLevel2':[company name, company code]}}
        """

        wb = load_workbook(filename=self.excel_file)
        ws = wb["Danh sách tất cả công ty"]
        stock_exchange_dict = dict()

        # to skip the heading
        g = ws.iter_rows(values_only=True)
        next(g)
        # -------------------#

        for row in g:
            try:
                (
                    _,
                    code_vs,
                    company_name,
                    industry_level_1,
                    industry_level_2,
                    exchange,
                    issue_amount,
                ) = row
            except ValueError:
                print(row)
                raise ValueError("this is the bad one")
            if exchange not in stock_exchange_dict:
                stock_exchange_dict[exchange] = {}
            if industry_level_2 not in stock_exchange_dict[exchange]:
                stock_exchange_dict[exchange][industry_level_2] = []
            stock_exchange_dict[exchange][industry_level_2].append(
                (
                    company_name,
                    code_vs,
                )
            )
        return stock_exchange_dict

    def company_to_db(self, Company, StockExchange, Industry):
        """
        input: Company model, StockExchange model, Industry model
        output: create object for Company model
        """

        for stock_exchange, value_1 in self.stock_exchange_dict.items():
            exchange = StockExchange.objects.get(stock_exchange=stock_exchange)
            for industry_level_2, value_2 in value_1.items():
                try:
                    industry = Industry.objects.get(industry_level_2=industry_level_2)
                    for value_3 in value_2:
                        name, code_vs = value_3
                        Company.objects.create(
                            code_vs=code_vs,
                            name=name,
                            industry=industry,
                            exchange=exchange,
                        )
                except ObjectDoesNotExist as e:
                    print(industry_level_2, value_2)
                    raise e


class MatchCode(object):
    def __init__(self, vs_file, fiin_file):
        self.vs_file = vs_file
        self.fiin_file = fiin_file
        self.fiin = self.get_fiin_df()
        self.vs = self.get_vs_df()
        self.fiin_dict = self.get_fiin_dict()
        self.vs_dict = self.get_vs_dict()
        self.fiin_diff, self.vs_diff, self.shared = self.get_diff()
        self.matches = self.match_effort()
        self.ticket_dict = self.prepare_ticket_dict()

    def get_fiin_df(self):
        fiin = pd.read_excel(
            self.fiin_file, header=8, index_col=0, sheet_name="ExportData"
        )
        try:
            assert fiin.columns[0] == "Ticker"
            assert type(fiin.index[-15]) == float
        except AssertionError as e:
            raise e(
                "The data structure is not as expected, please check the difference"
            )
        else:
            fiin.drop(fiin.tail(15).index, inplace=True)
            return fiin

    def get_vs_df(self):
        vs = pd.read_excel(
            self.vs_file, header=0, index_col=0, sheet_name="Danh sách tất cả công ty"
        )
        try:
            assert vs.columns[0] == "Mã"
        except AssertionError as e:
            raise e('Expect "Mã" as the first column header')
        else:
            return vs

    def get_fiin_dict(self):
        """
        output: dictionary {'Ticker': 'OrganShortName'}
        """

        df = self.fiin.copy()
        df.set_index("Ticker", inplace=True)
        return df.to_dict()["OrganShortName"]

    def get_vs_dict(self):
        """
        output: dictionary {'Mã': 'Tên công ty'}
        """

        df = self.vs.copy()
        df.set_index("Mã", inplace=True)
        return df.to_dict()["Tên công ty"]

    def get_diff(self):
        """
        output:
            fiin_diff: list of (company code, company name) for companies that are only in fiin dataset
            vs_diff: list of (company code, company name) for companies that are only in vs dataset
            shared: list of (company code, company name) for companies that are in both dataset
        """

        fiin_tickers = set(self.fiin_dict.keys())
        vs_tickers = set(self.vs_dict.keys())

        fiin_diff = [
            (ticker, self.fiin_dict[ticker]) for ticker in (fiin_tickers - vs_tickers)
        ]
        vs_diff = [
            (ticker, self.vs_dict[ticker]) for ticker in (vs_tickers - fiin_tickers)
        ]
        shared = [
            (ticker, self.fiin_dict[ticker]) for ticker in (vs_tickers & fiin_tickers)
        ]

        return (fiin_diff, vs_diff, shared)

    def match_effort(self):
        """
        output: dictionary of match fiin company name to vs company name
        """

        # Try to match the different companies from fiin to vs dataset
        matches = []
        for fiin_com in self.fiin_diff:
            for vs_com in self.vs_diff:
                q = re.compile(r".*{}.*".format(fiin_com[1]), re.IGNORECASE)
                m = q.match(vs_com[1])
                if m:
                    matches.append((fiin_com, vs_com))
        return matches

    def prepare_ticket_dict(self):
        """
        return the final dictionary: {vs_code: fiin_code}
        """
        ticket_dict = dict()
        for company in self.shared:
            ticket_dict[company[0]] = company[0]

        for match in self.matches:
            ticket_dict[match[1][0]] = match[0][0]

        return ticket_dict

    def add_fiin_code(self, Company):
        """
        input: Company model
        output: add fiin_code to company objects in Company model
        """

        for code_vs in self.ticket_dict:
            company = Company.objects.get(code_vs=code_vs)
            company.code_fiin = self.ticket_dict[code_vs]
            company.save()
